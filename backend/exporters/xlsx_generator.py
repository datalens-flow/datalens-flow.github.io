import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from backend.models import ExportRequest

def generate_xlsx_bytes(req: ExportRequest) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Sheet 1: Index of tables
    ws_index = wb.create_sheet(title="Tables Index")
    ws_index.views.sheetView[0].showGridLines = True
    
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    ws_index.append(["Table Name", "Schema", "Columns Count"])
    for cell in ws_index[1]:
        cell.font = header_font
        cell.fill = header_fill
        
    for table in req.schema_data.tables:
        ws_index.append([table.name, table.schema_name or "default", len(table.columns)])
        
    # Sheet 2: Unified Data Dictionary
    ws_dict = wb.create_sheet(title="Data Dictionary")
    ws_dict.views.sheetView[0].showGridLines = True
    ws_dict.append(["Table", "Column", "Type", "Nullable", "PK", "FK", "FK Reference", "Default", "Description"])
    for cell in ws_dict[1]:
        cell.font = header_font
        cell.fill = header_fill
        
    for table in req.schema_data.tables:
        table_desc = req.descriptions.get(table.id, {})
        for col in table.columns:
            fk_ref = f"{col.fk_ref_table}.{col.fk_ref_column}" if col.is_fk else ""
            desc = table_desc.get(col.name, col.comment or "")
            ws_dict.append([
                table.name,
                col.name,
                col.type,
                "Yes" if col.nullable else "No",
                "Yes" if col.is_pk else "No",
                "Yes" if col.is_fk else "No",
                fk_ref,
                col.default or "",
                desc
            ])
            
    for sheet in [ws_index, ws_dict]:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 10)
            for cell in col:
                cell.border = thin_border
                
    out = BytesIO()
    wb.save(out)
    return out.getvalue()
